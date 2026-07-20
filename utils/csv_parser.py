"""CSV and XLSX parser for attendee list validation in CertFlow."""

import csv
import io
import re
from typing import Dict, IO, List, Tuple, Union

from openpyxl import load_workbook

from utils.models import AttendeeRecord, ParseResult, ValidationError

EMAIL_REGEX = re.compile(r"^[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}$")


class CSVParser:
    """Reads and validates CSV files containing attendee name and email data.

    Handles header validation, row-level validation, duplicate detection,
    and returns structured results separating valid records from errors.
    """

    def parse(self, file: Union[IO[str], str]) -> ParseResult:
        """Parse a CSV file and return validated records and errors.

        Args:
            file: File-like object or string content of the CSV.

        Returns:
            ParseResult with valid records and validation errors.

        Raises:
            ValueError: If required column headers are missing.
        """
        if isinstance(file, str):
            file = io.StringIO(file)

        reader = csv.reader(file)

        # Read header row
        try:
            headers = next(reader)
        except StopIteration:
            raise ValueError("CSV file is empty — no header row found")

        # Validate headers
        column_map = self._validate_headers(headers)

        # Parse rows
        records: List[AttendeeRecord] = []
        errors: List[ValidationError] = []
        row_number = 1  # 1-based, header is row 0

        for row in reader:
            row_number += 1

            # Skip completely empty rows
            if not row or all(cell.strip() == "" for cell in row):
                continue

            name_idx = column_map["name"]
            email_idx = column_map["email"]

            # Extract values safely
            name = row[name_idx].strip() if name_idx < len(row) else ""
            email = row[email_idx].strip() if email_idx < len(row) else ""

            # Validate row
            record, row_errors = self._validate_row(row_number, name, email)
            if record:
                records.append(record)
            errors.extend(row_errors)

        # Detect duplicates
        records, duplicate_errors = self._detect_duplicates(records)
        errors.extend(duplicate_errors)

        return ParseResult(records=records, errors=errors)

    def _validate_headers(self, headers: List[str]) -> Dict[str, int]:
        """Validate required column headers exist (case-insensitive).

        Args:
            headers: List of header strings from the CSV first row.

        Returns:
            Dict mapping normalized header name to column index.

        Raises:
            ValueError: If 'name' or 'email' columns are missing.
        """
        normalized = [h.strip().lower() for h in headers]

        has_name = "name" in normalized
        has_email = "email" in normalized

        if not has_name and not has_email:
            raise ValueError(
                "CSV is missing required columns: 'name' and 'email'"
            )
        if not has_name:
            raise ValueError("CSV is missing required column: 'name'")
        if not has_email:
            raise ValueError("CSV is missing required column: 'email'")

        return {
            "name": normalized.index("name"),
            "email": normalized.index("email"),
        }

    def _validate_row(
        self, row_number: int, name: str, email: str
    ) -> Tuple:
        """Validate a single row's name and email fields.

        Args:
            row_number: 1-based row number for error reporting.
            name: Stripped name value.
            email: Stripped email value.

        Returns:
            Tuple of (AttendeeRecord or None, list of ValidationError).
        """
        row_errors: List[ValidationError] = []

        if not name:
            row_errors.append(
                ValidationError(
                    row_number=row_number,
                    field="name",
                    message="Name field is required (cannot be empty)",
                )
            )

        if not email:
            row_errors.append(
                ValidationError(
                    row_number=row_number,
                    field="email",
                    message="Email field is required (cannot be empty)",
                )
            )
        elif not EMAIL_REGEX.match(email):
            row_errors.append(
                ValidationError(
                    row_number=row_number,
                    field="email",
                    message=f"Invalid email format: '{email}'",
                )
            )

        if row_errors:
            return None, row_errors

        return AttendeeRecord(name=name, email=email), []

    def _detect_duplicates(
        self, records: List[AttendeeRecord]
    ) -> Tuple[List[AttendeeRecord], List[ValidationError]]:
        """Detect duplicate emails (case-insensitive), keeping first occurrence.

        Args:
            records: List of validated attendee records.

        Returns:
            Tuple of (deduplicated records, list of duplicate errors).
        """
        seen_emails: Dict[str, int] = {}
        deduplicated: List[AttendeeRecord] = []
        duplicate_errors: List[ValidationError] = []

        for i, record in enumerate(records):
            email_lower = record.email.lower()
            # header is row 1, data starts at row 2
            row_num = i + 2

            if email_lower in seen_emails:
                duplicate_errors.append(
                    ValidationError(
                        row_number=row_num,
                        field="duplicate",
                        message=(
                            f"Duplicate email '{record.email}' — "
                            f"first appeared at row {seen_emails[email_lower]}"
                        ),
                    )
                )
            else:
                seen_emails[email_lower] = row_num
                deduplicated.append(record)

        return deduplicated, duplicate_errors

    def parse_xlsx(self, file_bytes: bytes) -> ParseResult:
        """Parse an XLSX file and return validated records and errors.

        Args:
            file_bytes: Raw bytes of the .xlsx file.

        Returns:
            ParseResult with valid records and validation errors.

        Raises:
            ValueError: If required column headers are missing or file is invalid.
        """
        try:
            wb = load_workbook(filename=io.BytesIO(file_bytes), read_only=True)
        except Exception as e:
            raise ValueError(f"Cannot read XLSX file: {e}")

        ws = wb.active
        if ws is None:
            raise ValueError("XLSX file has no active worksheet")

        rows = list(ws.iter_rows(values_only=True))
        wb.close()

        if not rows:
            raise ValueError("XLSX file is empty — no header row found")

        # Validate headers
        headers = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
        column_map = self._validate_headers(headers)

        # Parse data rows
        records: List[AttendeeRecord] = []
        errors: List[ValidationError] = []

        for row_number, row in enumerate(rows[1:], start=2):
            # Skip completely empty rows
            if not row or all(
                cell is None or str(cell).strip() == "" for cell in row
            ):
                continue

            name_idx = column_map["name"]
            email_idx = column_map["email"]

            # Extract values safely
            name = (
                str(row[name_idx]).strip()
                if name_idx < len(row) and row[name_idx] is not None
                else ""
            )
            email = (
                str(row[email_idx]).strip()
                if email_idx < len(row) and row[email_idx] is not None
                else ""
            )

            # Validate row
            record, row_errors = self._validate_row(row_number, name, email)
            if record:
                records.append(record)
            errors.extend(row_errors)

        # Detect duplicates
        records, duplicate_errors = self._detect_duplicates(records)
        errors.extend(duplicate_errors)

        return ParseResult(records=records, errors=errors)

    def format_records(self, records: List[AttendeeRecord]) -> str:
        """Format attendee records back to CSV string.

        Args:
            records: List of AttendeeRecord objects.

        Returns:
            CSV string with 'name' and 'email' headers, rows in input order.
        """
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(["name", "email"])
        for record in records:
            writer.writerow([record.name, record.email])
        return output.getvalue()
